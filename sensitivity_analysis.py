# -*- coding: utf-8 -*-
"""
Kapsamlı Duyarlılık Analizi — Orman Yangını Su Tankı Yer Seçimi MILP
======================================================================
Analiz edilen parametreler:
  1. Talep miktarları   d_i   (±%10, ±%20, ±%30)
  2. Dış su maliyeti    c_ext (±%20, ±%40, ±%60, ±%100)
  3. Tank açılış maliy. F_k   (±%10, ±%20, ±%30)
  4. Tank işletme maliy O_k   (±%10, ±%20, ±%30)
  5. Tank kapasiteleri  C_k   (±%10, ±%20)
  6. Tornado analizi özeti

Çıktılar: sensitivity_results.xlsx  (7 sekme + tornado grafiği)
"""

from pathlib import Path
import copy
import pandas as pd
import pulp as pl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

INPUT_FILE  = Path("input.xlsx")
OUTPUT_FILE = Path("sensitivity_results.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(color="FFFFFF", bold=True, name="Arial", size=10)
UP_FILL  = PatternFill("solid", fgColor="C6EFCE")
DN_FILL  = PatternFill("solid", fgColor="FFC7CE")
NEU_FILL = PatternFill("solid", fgColor="FFEB9C")
BODY_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="BDD7EE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── Veri okuma ────────────────────────────────────────────────────────────────
def load_params():
    demand_df   = pd.read_excel(INPUT_FILE, sheet_name="DemandPoints")
    location_df = pd.read_excel(INPUT_FILE, sheet_name="CandidateLocations")
    tank_df     = pd.read_excel(INPUT_FILE, sheet_name="TankTypes")
    time_df     = pd.read_excel(INPUT_FILE, sheet_name="TravelTimes")
    config_df   = pd.read_excel(INPUT_FILE, sheet_name="Config")

    config = {str(r["Parameter"]): r["Value"] for _, r in config_df.iterrows()}

    I = demand_df["i"].astype(str).tolist()
    J = location_df["j"].astype(str).tolist()
    K = tank_df["k"].astype(str).tolist()

    d = dict(zip(demand_df["i"].astype(str), demand_df["d_i_m3"].astype(float)))
    T = dict(zip(demand_df["i"].astype(str), demand_df["T_i_min"].astype(float)))
    C = dict(zip(tank_df["k"].astype(str), tank_df["C_k_capacity_m3"].astype(float)))
    F = dict(zip(tank_df["k"].astype(str), tank_df["F_k_fixed_cost"].astype(float)))
    O = dict(zip(tank_df["k"].astype(str), tank_df["O_k_annual_cost"].astype(float)))
    u = dict(zip(location_df["j"].astype(str), location_df["u_j_protected"].astype(int)))
    S = dict(zip(location_df["j"].astype(str), location_df["S_j_slope_pct"].astype(float)))
    M = dict(zip(location_df["j"].astype(str), location_df["M_j_distance_m"].astype(float)))
    R = dict(zip(location_df["j"].astype(str), location_df["R_j_harvest_m3"].astype(float)))

    t = {}
    for _, row in time_df.iterrows():
        i = str(row["i"])
        for j in J:
            t[i, j] = float(row[j])

    a = {(i, j): (1 if t[i, j] <= T[i] else 0) for i in I for j in J}

    return dict(I=I, J=J, K=K, d=d, T=T, C=C, F=F, O=O,
                u=u, S=S, M=M, R=R, t=t, a=a,
                c_ext=float(config.get("c_ext", 85)),
                S_max=float(config.get("S_max", 25)),
                M_min=float(config.get("M_min", 250)))

# ── Model çözücü ─────────────────────────────────────────────────────────────
def solve(p):
    I, J, K = p["I"], p["J"], p["K"]
    d, C, F, O = p["d"], p["C"], p["F"], p["O"]
    u, S, M, R = p["u"], p["S"], p["M"], p["R"]
    a, c_ext = p["a"], p["c_ext"]
    S_max, M_min = p["S_max"], p["M_min"]
    total_demand = sum(d[i] for i in I)

    mdl = pl.LpProblem("SA", pl.LpMinimize)
    x = pl.LpVariable.dicts("x", (J, K), cat="Binary")
    y = pl.LpVariable.dicts("y", (I, J), cat="Binary")
    q = pl.LpVariable.dicts("q", (I, J), lowBound=0)
    w = pl.LpVariable.dicts("w", J, lowBound=0)

    mdl += (pl.lpSum(F[k]*x[j][k] for j in J for k in K)
          + pl.lpSum(O[k]*x[j][k] for j in J for k in K)
          + c_ext * pl.lpSum(w[j] for j in J))

    for i in I:
        for j in J:
            mdl += y[i][j] <= a[i,j]*pl.lpSum(x[j][k] for k in K)
    for i in I:
        mdl += pl.lpSum(y[i][j] for j in J) >= 1
    for i in I:
        mdl += pl.lpSum(q[i][j] for j in J) >= d[i]
    for i in I:
        for j in J:
            mdl += q[i][j] <= d[i]*y[i][j]
    for j in J:
        mdl += pl.lpSum(q[i][j] for i in I) <= pl.lpSum(C[k]*x[j][k] for k in K)
    for j in J:
        opened = pl.lpSum(x[j][k] for k in K)
        mdl += pl.lpSum(q[i][j] for i in I) <= R[j]*opened + w[j]
        mdl += w[j] <= total_demand*opened
    for j in J:
        mdl += pl.lpSum(x[j][k] for k in K) <= 1
    for j in J:
        mdl += pl.lpSum(x[j][k] for k in K) <= 1-int(u[j])
    for j in J:
        if S[j] > S_max:
            for k in K: mdl += x[j][k] == 0
    for j in J:
        if M[j] < M_min:
            for k in K: mdl += x[j][k] == 0

    mdl.solve(pl.PULP_CBC_CMD(msg=False))

    obj = pl.value(mdl.objective)
    _v = lambda var: (pl.value(var) or 0.0)
    opened = sum(1 for j in J for k in K if _v(x[j][k]) > 0.5)
    covered = sum(1 for i in I if sum(_v(y[i][j]) for j in J) >= 0.5)
    ext_water = sum(_v(w[j]) for j in J)
    selected = {j: next((k for k in K if _v(x[j][k]) > 0.5), None) for j in J}
    return dict(status=pl.LpStatus[mdl.status], obj=obj,
                opened=opened, covered=covered,
                ext_water=ext_water, selected=selected)

def pct(base, new):
    if base is None or base == 0 or new is None: return None
    return round((new-base)/abs(base)*100, 2)

def scale(d, f): return {k: v*f for k, v in d.items()}

# ── Çöz ──────────────────────────────────────────────────────────────────────
print("Parametreler yükleniyor...")
base_p = load_params()
print("Baz senaryo çözülüyor...")
base_r = solve(base_p)
B = base_r["obj"]
print(f"  Baz maliyet: {B:,.0f} TL | Durum: {base_r['status']}")

tables = {}

# 0 — Baz özet
tables["0_Baz_Senaryo"] = pd.DataFrame([
    ["Çözüm Durumu",            base_r["status"]],
    ["Toplam Maliyet (TL)",     round(B, 0)],
    ["Açılan Tank Sayısı",      base_r["opened"]],
    ["Kapsanan Talep Noktası",  base_r["covered"]],
    ["Dış Su İhtiyacı (m³)",    round(base_r["ext_water"],1)],
    ["c_ext (TL/m³)",           base_p["c_ext"]],
    ["S_max (%)",               base_p["S_max"]],
    ["M_min (m)",               base_p["M_min"]],
    ["Tank K1 Kap. (m³)",       base_p["C"]["K1"]],
    ["Tank K2 Kap. (m³)",       base_p["C"]["K2"]],
    ["Tank K3 Kap. (m³)",       base_p["C"]["K3"]],
], columns=["Metrik", "Baz Değer"])

# 1 — Talep
print("\n[1/5] Talep duyarlılığı (d_i)...")
rows = []
for p2ct in [-30,-20,-10,0,10,20,30]:
    p2 = copy.deepcopy(base_p); p2["d"] = scale(base_p["d"], 1+p2ct/100)
    r = solve(p2)
    rows.append({"Değişim (%)": p2ct, "Faktör": 1+p2ct/100,
                 "Toplam Maliyet (TL)": round(r["obj"],0),
                 "Maliyet Değişimi (%)": pct(B, r["obj"]),
                 "Açılan Tank": r["opened"],
                 "Kapsanan Nokta": r["covered"],
                 "Dış Su (m³)": round(r["ext_water"],1),
                 "Durum": r["status"]})
    print(f"  {p2ct:+d}% → {r['obj']:,.0f} TL")
tables["1_Talep_d_i"] = pd.DataFrame(rows)

# 2 — c_ext
print("\n[2/5] c_ext duyarlılığı...")
rows = []
for p2ct in [-40,-20,0,20,40,60,100]:
    p2 = copy.deepcopy(base_p); p2["c_ext"] = base_p["c_ext"]*(1+p2ct/100)
    r = solve(p2)
    rows.append({"Değişim (%)": p2ct, "c_ext (TL/m³)": round(p2["c_ext"],1),
                 "Toplam Maliyet (TL)": round(r["obj"],0),
                 "Maliyet Değişimi (%)": pct(B, r["obj"]),
                 "Açılan Tank": r["opened"],
                 "Dış Su (m³)": round(r["ext_water"],1),
                 "Durum": r["status"]})
    print(f"  {p2ct:+d}% → {r['obj']:,.0f} TL")
tables["2_Dis_Su_Maliyeti_cext"] = pd.DataFrame(rows)

# 3 — F_k
print("\n[3/5] Açılış maliyeti duyarlılığı (F_k)...")
rows = []
for p2ct in [-30,-20,-10,0,10,20,30]:
    p2 = copy.deepcopy(base_p); p2["F"] = scale(base_p["F"], 1+p2ct/100)
    r = solve(p2)
    rows.append({"Değişim (%)": p2ct,
                 "F_K1 (TL)": round(p2["F"]["K1"],0),
                 "F_K2 (TL)": round(p2["F"]["K2"],0),
                 "F_K3 (TL)": round(p2["F"]["K3"],0),
                 "Toplam Maliyet (TL)": round(r["obj"],0),
                 "Maliyet Değişimi (%)": pct(B, r["obj"]),
                 "Açılan Tank": r["opened"],
                 "Durum": r["status"]})
    print(f"  {p2ct:+d}% → {r['obj']:,.0f} TL")
tables["3_Acilis_Maliyeti_Fk"] = pd.DataFrame(rows)

# 4 — O_k
print("\n[4/5] İşletme maliyeti duyarlılığı (O_k)...")
rows = []
for p2ct in [-30,-20,-10,0,10,20,30]:
    p2 = copy.deepcopy(base_p); p2["O"] = scale(base_p["O"], 1+p2ct/100)
    r = solve(p2)
    rows.append({"Değişim (%)": p2ct,
                 "O_K1 (TL/yıl)": round(p2["O"]["K1"],0),
                 "O_K2 (TL/yıl)": round(p2["O"]["K2"],0),
                 "O_K3 (TL/yıl)": round(p2["O"]["K3"],0),
                 "Toplam Maliyet (TL)": round(r["obj"],0),
                 "Maliyet Değişimi (%)": pct(B, r["obj"]),
                 "Açılan Tank": r["opened"],
                 "Durum": r["status"]})
    print(f"  {p2ct:+d}% → {r['obj']:,.0f} TL")
tables["4_Isletme_Maliyeti_Ok"] = pd.DataFrame(rows)

# 5 — C_k
print("\n[5/5] Tank kapasitesi duyarlılığı (C_k)...")
rows = []
for p2ct in [-20,-10,0,10,20]:
    p2 = copy.deepcopy(base_p); p2["C"] = scale(base_p["C"], 1+p2ct/100)
    r = solve(p2)
    rows.append({"Değişim (%)": p2ct,
                 "C_K1 (m³)": round(p2["C"]["K1"],0),
                 "C_K2 (m³)": round(p2["C"]["K2"],0),
                 "C_K3 (m³)": round(p2["C"]["K3"],0),
                 "Toplam Maliyet (TL)": round(r["obj"],0),
                 "Maliyet Değişimi (%)": pct(B, r["obj"]),
                 "Açılan Tank": r["opened"],
                 "Dış Su (m³)": round(r["ext_water"],1),
                 "Durum": r["status"]})
    print(f"  {p2ct:+d}% → {r['obj']:,.0f} TL")
tables["5_Kapasite_Ck"] = pd.DataFrame(rows)

# 6 — Tornado
print("\nTornado analizi...")
def tornado(label, lo_pct, hi_pct, lo_p, hi_p):
    lo_r = solve(lo_p); hi_r = solve(hi_p)
    return {"Parametre": label,
            "Alt Senaryo (%)": lo_pct, "Üst Senaryo (%)": hi_pct,
            "Alt Maliyet (TL)": round(lo_r["obj"],0),
            "Baz Maliyet (TL)": round(B,0),
            "Üst Maliyet (TL)": round(hi_r["obj"],0),
            "Maliyet Aralığı (TL)": round(abs(hi_r["obj"]-lo_r["obj"]),0),
            "Alt Δ (%)": pct(B, lo_r["obj"]),
            "Üst Δ (%)": pct(B, hi_r["obj"])}

t_rows = []
p_lo = copy.deepcopy(base_p); p_lo["d"] = scale(base_p["d"],0.80)
p_hi = copy.deepcopy(base_p); p_hi["d"] = scale(base_p["d"],1.20)
t_rows.append(tornado("Talep d_i", -20, +20, p_lo, p_hi))

p_lo = copy.deepcopy(base_p); p_lo["c_ext"] = base_p["c_ext"]*0.60
p_hi = copy.deepcopy(base_p); p_hi["c_ext"] = base_p["c_ext"]*1.40
t_rows.append(tornado("Dış Su Maliyeti c_ext", -40, +40, p_lo, p_hi))

p_lo = copy.deepcopy(base_p); p_lo["F"] = scale(base_p["F"],0.70)
p_hi = copy.deepcopy(base_p); p_hi["F"] = scale(base_p["F"],1.30)
t_rows.append(tornado("Açılış Maliyeti F_k", -30, +30, p_lo, p_hi))

p_lo = copy.deepcopy(base_p); p_lo["O"] = scale(base_p["O"],0.70)
p_hi = copy.deepcopy(base_p); p_hi["O"] = scale(base_p["O"],1.30)
t_rows.append(tornado("İşletme Maliyeti O_k", -30, +30, p_lo, p_hi))

p_lo = copy.deepcopy(base_p); p_lo["C"] = scale(base_p["C"],0.80)
p_hi = copy.deepcopy(base_p); p_hi["C"] = scale(base_p["C"],1.20)
t_rows.append(tornado("Tank Kapasitesi C_k", -20, +20, p_lo, p_hi))

tornado_df = pd.DataFrame(t_rows).sort_values("Maliyet Aralığı (TL)", ascending=False)
tables["6_Tornado_Ozet"] = tornado_df

# ── Excel yaz ─────────────────────────────────────────────────────────────────
print("\nExcel dosyası yazılıyor...")
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    for sname, df in tables.items():
        df.to_excel(writer, sheet_name=sname[:31], index=False)

wb2 = load_workbook(OUTPUT_FILE)

for ws in wb2.worksheets:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
        alt = PatternFill("solid", fgColor="EBF3FB") if ri % 2 == 0 else PatternFill()
        for cell in row:
            cell.border = BORDER; cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            hdr = ws.cell(row=1, column=cell.column).value or ""
            if "Değişim" in str(hdr) and "%" in str(hdr) and isinstance(cell.value, (int,float)):
                cell.fill = UP_FILL if cell.value < -0.1 else (DN_FILL if cell.value > 0.1 else NEU_FILL)
                cell.number_format = "+0.00;-0.00;0.00"
            elif "TL" in str(hdr) and isinstance(cell.value, (int,float)):
                cell.fill = alt; cell.number_format = "#,##0"
            else:
                cell.fill = alt
    for ci, col_cells in enumerate(ws.columns, 1):
        mlen = max((len(str(c.value)) for c in col_cells if c.value), default=8)
        ws.column_dimensions[get_column_letter(ci)].width = min(mlen+4, 40)
    ws.row_dimensions[1].height = 30

# Tornado grafiği
ws_t = wb2["6_Tornado_Ozet"]
chart = BarChart()
chart.type = "bar"
chart.title = "Tornado Analizi — Maliyet Üzerindeki Etki Büyüklüğü"
chart.y_axis.title = "Parametre"
chart.x_axis.title = "Maliyet Aralığı (TL)"
chart.style = 10; chart.width = 24; chart.height = 14
data_ref = Reference(ws_t, min_col=7, min_row=1, max_row=ws_t.max_row)
cats_ref = Reference(ws_t, min_col=1, min_row=2, max_row=ws_t.max_row)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.series[0].graphicalProperties.solidFill = "2E75B6"
ws_t.add_chart(chart, "J2")

wb2.save(OUTPUT_FILE)
print(f"\n{'='*55}")
print(f"  TAMAMLANDI → {OUTPUT_FILE}")
print(f"{'='*55}")
print(f"  Baz maliyet  : {B:,.0f} TL")
print(f"  Açılan tank  : {base_r['opened']}")
print(f"  Kapsama      : {base_r['covered']}/{len(base_p['I'])} nokta")
print(f"  Dış su       : {base_r['ext_water']:.1f} m³")
print(f"\nSekme listesi:")
for s in wb2.sheetnames:
    print(f"  • {s}")
