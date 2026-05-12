# -*- coding: utf-8 -*-
"""
Rainwater-Integrated Water Tank Location MILP Model
Rapora birebir uyumlu - DÜZELTİLMİŞ VERSİYON

Düzeltilen hatalar:
  1. KISIT 6: Sadece tank kapasitesi Σ C_k x_jk (yağmur R_j ayrı kısıtta)
  2. KISIT 6b: Kapasite dengesi R_j ve w_j ile genişletme doğru yerde
  3. KISIT 11 (min_served): Config'de yoksa eklenmez (rapora fazladan kısıt)
  4. Amaç fonksiyonu: Değişmedi, zaten doğruydu

Kurulum:
    pip install pandas openpyxl pulp

Çalıştırma:
    python water_tank_pulp_model_FIXED.py
"""

from pathlib import Path
import math
import pandas as pd

try:
    import pulp as pl
except ModuleNotFoundError as exc:
    raise ModuleNotFoundError("PuLP kurulu değil. pip install pulp") from exc

BASE_DIR    = Path(__file__).resolve().parent
INPUT_FILE  = BASE_DIR / "input.xlsx"
OUTPUT_FILE = BASE_DIR / "water_tank_solution_FIXED.xlsx"


# =============================================================================
# 1) EXCEL OKUMA VE PARAMETRE HAZIRLAMA
# =============================================================================

def read_config(config_df: pd.DataFrame) -> dict:
    config = {}
    for _, row in config_df.iterrows():
        config[str(row["Parameter"])] = row["Value"]
    return config


def build_parameters() -> dict:
    demand_df   = pd.read_excel(INPUT_FILE, sheet_name="DemandPoints")
    location_df = pd.read_excel(INPUT_FILE, sheet_name="CandidateLocations")
    tank_df     = pd.read_excel(INPUT_FILE, sheet_name="TankTypes")
    time_df     = pd.read_excel(INPUT_FILE, sheet_name="TravelTimes")
    config_df   = pd.read_excel(INPUT_FILE, sheet_name="Config")

    config = read_config(config_df)

    I = demand_df["i"].astype(str).tolist()
    J = location_df["j"].astype(str).tolist()
    K = tank_df["k"].astype(str).tolist()

    d = dict(zip(demand_df["i"].astype(str), demand_df["d_i_m3"].astype(float)))
    T = dict(zip(demand_df["i"].astype(str), demand_df["T_i_min"].astype(float)))

    C = dict(zip(tank_df["k"].astype(str), tank_df["C_k_capacity_m3"].astype(float)))
    F = dict(zip(tank_df["k"].astype(str), tank_df["F_k_fixed_cost"].astype(float)))
    O = dict(zip(tank_df["k"].astype(str), tank_df["O_k_annual_cost"].astype(float)))

    u = dict(zip(location_df["j"].astype(str), location_df["u_j_protected"].astype(int)))
    S = dict(zip(location_df["j"].astype(str), location_df["S_j_slope_pct"].astype(float)))
    M = dict(zip(location_df["j"].astype(str), location_df["M_j_distance_m"].astype(float)))

    c_ext = float(config.get("c_ext", 70))
    S_max = float(config.get("S_max", 18))
    M_min = float(config.get("M_min", 250))

    # DÜZELTİLDİ: min_served sadece config'de varsa kullanılır
    min_served_raw = config.get("min_served_demand_per_open_tank", None)
    min_served = int(float(min_served_raw)) if min_served_raw is not None else None

    # Yağmur suyu hasadı: R_j = P_j * A_j * phi_j * eta_j / 1000
    R = {}
    for _, row in location_df.iterrows():
        j = str(row["j"])
        if "R_j_harvest_m3" in location_df.columns and pd.notna(row.get("R_j_harvest_m3")):
            R[j] = float(row["R_j_harvest_m3"])
        else:
            R[j] = (
                float(row["P_j_rainfall_mm"])
                * float(row["A_j_collection_area_m2"])
                * float(row["phi_j_runoff"])
                * float(row["eta_j_efficiency"])
                / 1000.0
            )

    t = {}
    for _, row in time_df.iterrows():
        i = str(row["i"])
        for j in J:
            t[i, j] = float(row[j])

    # a_ij = 1 if t_ij <= T_i, 0 otherwise
    a = {(i, j): (1 if t[i, j] <= T[i] else 0) for i in I for j in J}

    return {
        "I": I, "J": J, "K": K,
        "d": d, "T": T,
        "C": C, "F": F, "O": O,
        "u": u, "S": S, "M": M,
        "R": R, "t": t, "a": a,
        "c_ext": c_ext, "S_max": S_max, "M_min": M_min,
        "min_served": min_served,
        "demand_df": demand_df,
        "location_df": location_df,
        "tank_df": tank_df,
        "config": config,
    }


# =============================================================================
# 2) MILP MODELİ — RAPORA BİREBİR UYUMLU (DÜZELTİLMİŞ)
# =============================================================================

def solve_model(params: dict):
    """
    Rapordaki matematiksel modeli PuLP ile kurar.

    AMAÇ FONKSİYONU (Denklem 1):
        min Z = ΣΣ F_k x_jk  +  ΣΣ O_k x_jk  +  c_ext Σ w_j

    KISITLAR:
        (2)  y_ij <= a_ij * Σ_k x_jk                    ∀i,j
        (3)  Σ_j y_ij >= 1                               ∀i ∈ I
        (4)  Σ_j q_ij >= d_i                             ∀i ∈ I
        (5)  q_ij <= d_i * y_ij                          ∀i,j
        (6)  Σ_i q_ij <= Σ_k C_k x_jk                   ∀j   ← RAPOR: SADECE TANK KAPASİTESİ
        (6b) [Su dengesi] Σ_i q_ij <= R_j*opened_j + w_j ∀j   ← EKLEME: yağmur+dış kaynak dengeleme
             NOT: (6) ve (6b) birlikte uygulanır; efektif limit min(C_k, R_j+w_j)
        (7)  Σ_k x_jk <= 1                               ∀j
        (8)  Σ_k x_jk <= 1 - u_j                         ∀j
        (9)  x_jk = 0 if S_j > S_max                    ∀j,k
        (10) x_jk = 0 if M_j < M_min                    ∀j,k
        (11) [Opsiyonel] Σ_i y_ij >= min_served*opened_j ∀j  (sadece config'de tanımlıysa)

    DÜZELTİLEN TEMEL HATA:
        Önceki kodda kısıt 6: Σ_i q_ij <= Σ_k C_k x_jk + R_j*opened_j + w_j
        Bu yanlıştı: Tank kapasitesinin ÜZERİNE yağmur ve dış su ekliyordu,
        dolayısıyla 80m3'lük tanktan 456m3 su gönderilebiliyordu!

        Doğrusu: İki ayrı kısıt:
          - (6):  Σ_i q_ij <= Σ_k C_k x_jk       (tank doluluk sınırı)
          - (6b): Σ_i q_ij <= R_j*opened_j + w_j  (temin edilen su ≥ gönderilen su)
          + w_j kısıtı: w_j <= total_demand * opened_j (kapalı tankta dış su yok)
    """

    I, J, K  = params["I"], params["J"], params["K"]
    d, C, F, O = params["d"], params["C"], params["F"], params["O"]
    u, S, M, R = params["u"], params["S"], params["M"], params["R"]
    a           = params["a"]
    c_ext       = params["c_ext"]
    S_max, M_min = params["S_max"], params["M_min"]
    min_served  = params["min_served"]  # None veya int
    total_demand = sum(d[i] for i in I)

    model = pl.LpProblem("Rainwater_Integrated_Water_Tank_Location_MILP_FIXED", pl.LpMinimize)

    # Karar Değişkenleri
    x = pl.LpVariable.dicts("x", (J, K), cat="Binary")
    y = pl.LpVariable.dicts("y", (I, J), cat="Binary")
    q = pl.LpVariable.dicts("q", (I, J), lowBound=0, cat="Continuous")
    w = pl.LpVariable.dicts("w",  J,      lowBound=0, cat="Continuous")

    # Amaç Fonksiyonu (Denklem 1)
    model += (
        pl.lpSum(F[k] * x[j][k] for j in J for k in K)
        + pl.lpSum(O[k] * x[j][k] for j in J for k in K)
        + c_ext * pl.lpSum(w[j] for j in J)
    ), "Objective_Total_Cost"

    # Kısıt (2): Sadece erişilebilir ve açık tanklara atama
    for i in I:
        for j in J:
            model += (
                y[i][j] <= a[i, j] * pl.lpSum(x[j][k] for k in K)
            ), f"C02_{i}_{j}"

    # Kısıt (3): Her talep noktası en az bir tanka atanmalı
    for i in I:
        model += pl.lpSum(y[i][j] for j in J) >= 1, f"C03_{i}"

    # Kısıt (4): Her talep noktasının su ihtiyacı tamamen karşılanmalı
    for i in I:
        model += pl.lpSum(q[i][j] for j in J) >= d[i], f"C04_{i}"

    # Kısıt (5): Atama yoksa su akışı olamaz
    for i in I:
        for j in J:
            model += q[i][j] <= d[i] * y[i][j], f"C05_{i}_{j}"

    # *** DÜZELTİLDİ ***
    # Kısıt (6): Tank kapasitesi üst sınırı — RAPOR DENKLEMİ: Σ_i q_ij <= Σ_k C_k x_jk
    # Bu kısıt tankın fiziksel kapasitesini aşmamayı garantiler.
    for j in J:
        model += (
            pl.lpSum(q[i][j] for i in I) <= pl.lpSum(C[k] * x[j][k] for k in K)
        ), f"C06_tank_capacity_{j}"

    # Kısıt (6b): Su temin dengesi — yağmur hasadı + dış kaynak su >= gönderilen su
    # Rapordaki Ck = R_j * x_jk + w_j formülüne karşılık gelir.
    # Açık olmayan lokasyonda w_j ve R_j*opened_j sıfır olmalı.
    for j in J:
        opened_j = pl.lpSum(x[j][k] for k in K)
        model += (
            pl.lpSum(q[i][j] for i in I) <= R[j] * opened_j + w[j]
        ), f"C06b_water_supply_{j}"
        # Dış kaynak suyu yalnızca açık tankta anlam taşır
        model += w[j] <= total_demand * opened_j, f"C06c_ext_water_only_open_{j}"

    # Kısıt (7): Her lokasyonda en fazla bir tank tipi
    for j in J:
        model += pl.lpSum(x[j][k] for k in K) <= 1, f"C07_{j}"

    # Kısıt (8): Koruma altındaki lokasyonda tank açılamaz
    for j in J:
        model += (
            pl.lpSum(x[j][k] for k in K) <= 1 - int(u[j])
        ), f"C08_{j}"

    # Kısıt (9): Eğim fazlaysa tank açılamaz
    for j in J:
        if S[j] > S_max:
            for k in K:
                model += x[j][k] == 0, f"C09_{j}_{k}"

    # Kısıt (10): Yerleşime çok yakınsa tank açılamaz
    for j in J:
        if M[j] < M_min:
            for k in K:
                model += x[j][k] == 0, f"C10_{j}_{k}"

    # *** DÜZELTİLDİ ***
    # Kısıt (11): min_served — SADECE config'de tanımlıysa eklenir
    if min_served is not None and min_served > 0:
        for j in J:
            opened_j = pl.lpSum(x[j][k] for k in K)
            model += (
                pl.lpSum(y[i][j] for i in I) >= min_served * opened_j
            ), f"C11_{j}"
        print(f"[INFO] Kısıt (11) eklendi: min_served = {min_served}")
    else:
        print("[INFO] Kısıt (11) atlandı: Config'de min_served_demand_per_open_tank tanımlı değil.")

    solver = pl.PULP_CBC_CMD(msg=True)
    model.solve(solver)
    return model, x, y, q, w


# =============================================================================
# 3) SONUÇ TABLOLARI
# =============================================================================

def _val(var, default=0.0):
    v = pl.value(var)
    return default if (v is None or (isinstance(v, float) and math.isnan(v))) else float(v)


def _selected_tank(j, K, x):
    for k in K:
        if _val(x[j][k]) > 0.5:
            return k
    return ""


def build_report_tables(params, model, x, y, q, w):
    I, J, K = params["I"], params["J"], params["K"]
    d, T    = params["d"], params["T"]
    C, F, O = params["C"], params["F"], params["O"]
    R, t, a = params["R"], params["t"], params["a"]
    u, S, M = params["u"], params["S"], params["M"]
    c_ext, S_max, M_min = params["c_ext"], params["S_max"], params["M_min"]
    demand_df   = params["demand_df"]
    location_df = params["location_df"]
    min_served  = params["min_served"]

    status    = pl.LpStatus[model.status]
    objective = pl.value(model.objective)

    demand_lut   = demand_df.set_index("i")
    location_lut = location_df.set_index("j")

    fixed_cost_val     = sum(F[k] * _val(x[j][k]) for j in J for k in K)
    operating_cost_val = sum(O[k] * _val(x[j][k]) for j in J for k in K)
    ext_water_total    = sum(_val(w[j]) for j in J)
    ext_cost_val       = c_ext * ext_water_total

    opened_count     = sum(1 for j in J if _selected_tank(j, K, x))
    covered_count    = sum(1 for i in I if sum(_val(y[i][j]) for j in J) >= 0.5)
    total_demand_sum = sum(d[i] for i in I)
    total_sent       = sum(_val(q[i][j]) for i in I for j in J)
    rain_at_opened   = sum(R[j] for j in J if _selected_tank(j, K, x))
    ext_reduction    = (
        max(0.0, 1.0 - ext_water_total / total_sent)
        if total_sent > 1e-9 else 0.0
    )

    # Kapasite ihlali kontrolü (doğrulama)
    capacity_violations = []
    for j in J:
        k = _selected_tank(j, K, x)
        if k:
            sent_j = sum(_val(q[i][j]) for i in I)
            cap_j  = C[k]
            if sent_j > cap_j + 1e-3:
                capacity_violations.append(f"{j}(gönderilen={sent_j:.1f} > kapasite={cap_j})")

    summary_df = pd.DataFrame([
        ["Solver Status",                          status,                              "Modelin çözüm durumu"],
        ["Objective Total Cost (TL)",              objective,                           "Açılış + işletme + dış kaynak su maliyeti"],
        ["Fixed Opening Cost (TL)",                fixed_cost_val,                      "Seçilen tank tiplerinin toplam açılış maliyeti"],
        ["Annual Operating Cost (TL)",             operating_cost_val,                  "Seçilen tank tiplerinin yıllık işletme maliyeti"],
        ["External Water Cost (TL)",               ext_cost_val,                        "Dış kaynaktan taşınan su maliyeti (c_ext * Σ w_j)"],
        ["Opened Tank Count",                      opened_count,                        "Açılan toplam tank / lokasyon sayısı"],
        ["Covered Demand Point Count",             covered_count,                       "Ataması gerçekleşen talep / risk noktası sayısı"],
        ["Coverage Ratio",                         covered_count / len(I) if I else 0,  "Kapsama oranı (covered / total)"],
        ["Total Demand (m³)",                      total_demand_sum,                    "Tüm talep noktalarının toplam su ihtiyacı"],
        ["Total Water Sent (m³)",                  total_sent,                          "Tanklardan gönderilen toplam su"],
        ["Rainwater at Opened Tanks (m³/year)",    rain_at_opened,                      "Açılan lokasyonlardaki yıllık yağmur hasadı"],
        ["External Water Used (m³)",               ext_water_total,                     "Dış kaynak suyu ihtiyacı (Σ w_j)"],
        ["External Water Reduction Ratio",         ext_reduction,                       "Yağmur hasadı sayesinde dış su talebindeki azalma"],
        ["Capacity Violations (must be 0)",        len(capacity_violations),            "KAPASİTE İHLALİ (0 olmalı) - Doğrulama"],
        ["Constraint (11) min_served active",      str(min_served) if min_served else "Hayır (config'de yok)",
                                                                                        "Rapor kısıt 11 aktif mi?"],
    ], columns=["Metric", "Value", "Explanation"])

    # Optimal tank lokasyonları
    opened_rows = []
    for j in J:
        k = _selected_tank(j, K, x)
        if not k:
            continue
        loc = location_lut.loc[j]
        served = [i for i in I if _val(y[i][j]) > 0.5]
        sent_j = sum(_val(q[i][j]) for i in I)
        cap_j  = C[k]
        opened_rows.append({
            "j":                             j,
            "Region":                        loc.get("Region",              ""),
            "Province":                      loc.get("Province",            ""),
            "Candidate_Location":            loc.get("Candidate_Location",  ""),
            "Latitude":                      loc.get("Latitude",            ""),
            "Longitude":                     loc.get("Longitude",           ""),
            "Selected_Tank_Type_k":          k,
            "Tank_Capacity_C_k_m3":          cap_j,
            "Fixed_Cost_F_k":                F[k],
            "Annual_Operating_Cost_O_k":     O[k],
            "Rainwater_Harvest_R_j_m3":      R[j],
            "External_Water_w_j_m3":         _val(w[j]),
            "Total_Water_Sent_m3":           sent_j,
            "Capacity_Utilization_Pct":      round(sent_j / cap_j * 100, 1) if cap_j > 0 else 0,
            "Capacity_OK":                   "YES" if sent_j <= cap_j + 1e-3 else "VIOLATION!",
            "Served_Demand_Point_Count":     len(served),
            "Served_Demand_Points":          ", ".join(served),
            "Protected_Area_u_j":            u[j],
            "Slope_S_j_pct":                 S[j],
            "Slope_Limit_S_max":             S_max,
            "Distance_M_j_m":               M[j],
            "Min_Safety_Distance_M_min":     M_min,
            "Site_Eligibility":              "Suitable",
            "Detailed_Note":                 loc.get("Detailed_Note", ""),
        })
    opened_df = pd.DataFrame(opened_rows)

    # Talep noktası kapsama durumu
    coverage_rows = []
    for i in I:
        dem = demand_lut.loc[i]
        assigned = [j for j in J if _val(y[i][j]) > 0.5]
        sent_i   = sum(_val(q[i][j]) for j in J)
        min_time = min((t[i, j] for j in assigned), default=None)
        coverage_rows.append({
            "i":                                    i,
            "Region":                               dem.get("Region",       ""),
            "Province":                             dem.get("Province",     ""),
            "Demand_Point":                         dem.get("Demand_Point", ""),
            "Priority":                             dem.get("Priority",     ""),
            "Latitude":                             dem.get("Latitude",     ""),
            "Longitude":                            dem.get("Longitude",    ""),
            "Demand_d_i_m3":                        d[i],
            "Max_Response_Time_T_i_min":            T[i],
            "Assigned_Tank_Count":                  len(assigned),
            "Assigned_Tanks_j":                     ", ".join(assigned),
            "Min_Response_Time_from_Assigned_min":  min_time,
            "Total_Water_Received_m3":              sent_i,
            "Demand_Fulfilled":                     "YES" if sent_i + 1e-6 >= d[i] else "NO",
            "Coverage_Status":                      "Covered" if assigned else "NOT COVERED",
        })
    coverage_df = pd.DataFrame(coverage_rows)

    # Su akış matrisi
    flow_rows = []
    for i in I:
        for j in J:
            qv = _val(q[i][j])
            if qv > 1e-6:
                dem = demand_lut.loc[i]
                loc = location_lut.loc[j]
                flow_rows.append({
                    "From_Tank_j":               j,
                    "Tank_Location":             loc.get("Candidate_Location", ""),
                    "To_Demand_i":               i,
                    "Demand_Point":              dem.get("Demand_Point", ""),
                    "Travel_Time_t_ij_min":      t[i, j],
                    "Max_Response_Time_T_i_min": T[i],
                    "Accessible_a_ij":           a[i, j],
                    "Water_Flow_q_ij_m3":        qv,
                })
    flow_df = pd.DataFrame(flow_rows)

    # Lokasyon uygunluk değerlendirmesi
    feas_rows = []
    for j in J:
        loc = location_lut.loc[j]
        k   = _selected_tank(j, K, x)
        slope_ok   = S[j] <= S_max
        dist_ok    = M[j] >= M_min
        protect_ok = int(u[j]) == 0
        suitable   = protect_ok and slope_ok and dist_ok
        feas_rows.append({
            "j":                          j,
            "Region":                     loc.get("Region",             ""),
            "Province":                   loc.get("Province",           ""),
            "Candidate_Location":         loc.get("Candidate_Location", ""),
            "Opened":                     "YES" if k else "NO",
            "Selected_Tank_Type":         k,
            "Protected_Area_OK_u_j=0":   "YES" if protect_ok else "NO",
            "Slope_OK_S_j<=S_max":       "YES" if slope_ok   else "NO",
            "Distance_OK_M_j>=M_min":    "YES" if dist_ok    else "NO",
            "Overall_Site_Suitability":  "Suitable" if suitable else "Not Suitable",
            "u_j_protected":             u[j],
            "S_j_slope_pct":             S[j],
            "S_max":                     S_max,
            "M_j_distance_m":            M[j],
            "M_min":                     M_min,
            "Annual_Rainwater_R_j_m3":   R[j],
        })
    feas_df = pd.DataFrame(feas_rows)

    # Maliyet kırılımı
    cost_df = pd.DataFrame([
        ["Fixed Opening Cost",    fixed_cost_val,     "Σ_j Σ_k F_k · x_jk"],
        ["Annual Operating Cost", operating_cost_val, "Σ_j Σ_k O_k · x_jk"],
        ["External Water Cost",   ext_cost_val,       "c_ext · Σ_j w_j"],
        ["Total Cost",            objective,           "Objective function value"],
    ], columns=["Cost_Component", "Value_TL", "Formula_Reference"])

    # Kısıt referans sayfası (DÜZELTİLMİŞ)
    constraints_df = pd.DataFrame([
        ["(1)",  "Objective",              "min Σ F_k x_jk + Σ O_k x_jk + c_ext Σ w_j",
                 "Açılış + işletme + dış kaynak su maliyetini minimize et"],
        ["(2)",  "Assignment feasibility", "y_ij <= a_ij · Σ_k x_jk  ∀i,j",
                 "Sadece erişilebilir ve açık tanklara atama"],
        ["(3)",  "Full coverage",          "Σ_j y_ij >= 1  ∀i",
                 "Her talep noktası en az bir tanka atanmalı"],
        ["(4)",  "Demand fulfillment",     "Σ_j q_ij >= d_i  ∀i",
                 "Her talep noktasının su ihtiyacı tam karşılanmalı"],
        ["(5)",  "Flow-assignment link",   "q_ij <= d_i · y_ij  ∀i,j",
                 "Atama yoksa su akışı olamaz"],
        ["(6)",  "Tank capacity",          "Σ_i q_ij <= Σ_k C_k x_jk  ∀j",
                 "RAPOR DENKLEMİ: Sadece tank fiziksel kapasitesi (DÜZELTİLDİ)"],
        ["(6b)", "Water supply balance",   "Σ_i q_ij <= R_j·opened_j + w_j  ∀j",
                 "Yağmur hasadı + dış kaynak su dengesi (AYRI KISIT)"],
        ["(7)",  "One tank type per site", "Σ_k x_jk <= 1  ∀j",
                 "Her aday lokasyonda en fazla bir tank tipi"],
        ["(8)",  "Protected areas",        "Σ_k x_jk <= 1 - u_j  ∀j",
                 "Koruma alanında tank açılamaz"],
        ["(9)",  "Slope constraint",       "x_jk = 0  if S_j > S_max  ∀j,k",
                 "Eğim limitini aşan lokasyona tank açılamaz"],
        ["(10)", "Safety distance",        "x_jk = 0  if M_j < M_min  ∀j,k",
                 "Yerleşime çok yakın lokasyona tank açılamaz"],
        ["(11)", "Min utilization (opt)",  "Σ_i y_ij >= min_served · opened_j  ∀j  [Config'de varsa]",
                 "DÜZELTİLDİ: Sadece config'de tanımlıysa aktif olur"],
    ], columns=["Eq_No", "Constraint_Name", "Mathematical_Form", "Interpretation"])

    # Hata karşılaştırma özeti (eski vs yeni)
    error_summary_df = pd.DataFrame([
        ["KISIT 6 (Ana Hata)",
         "Σ_i q_ij <= Σ_k C_k x_jk + R_j*opened_j + w_j",
         "Σ_i q_ij <= Σ_k C_k x_jk  (ayrıca 6b ile su dengesi)",
         "Tank kapasitesi aşılıyordu! 80m3 tanktan 456m3 su gönderiliyordu."],
        ["KISIT 11 (min_served)",
         "Her zaman aktif (varsayılan min_served=1)",
         "Sadece config'de tanımlıysa aktif",
         "Raporda olmayan fazladan kısıt modeli kısıtlıyordu."],
        ["Kısıt 6b (Yeni)",
         "Yoktu",
         "Σ_i q_ij <= R_j*opened_j + w_j eklendi",
         "Su temin dengesi artık doğru modellendi."],
    ], columns=["Bileşen", "Eski (Hatalı)", "Yeni (Düzeltilmiş)", "Açıklama"])

    return {
        "Model_Summary":         summary_df,
        "Optimal_Tank_Locations": opened_df,
        "Demand_Coverage":       coverage_df,
        "Water_Flows":           flow_df,
        "Location_Feasibility":  feas_df,
        "Cost_Breakdown":        cost_df,
        "Constraint_Reference":  constraints_df,
        "Error_Fix_Summary":     error_summary_df,
    }


# =============================================================================
# 4) EXCEL YAZMA VE FORMATLAMA
# =============================================================================

def format_workbook(output_file: Path):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = load_workbook(output_file)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, name="Arial", size=10)
    alt_fill    = PatternFill("solid", fgColor="EBF3FB")
    body_font   = Font(name="Arial", size=10)
    warn_fill   = PatternFill("solid", fgColor="FF6B6B")
    warn_font   = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    thin        = Side(style="thin", color="BDD7EE")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    from openpyxl.utils import get_column_letter

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.border    = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = alt_fill if row_idx % 2 == 0 else PatternFill()
            for cell in row:
                # Hata/ihlal vurgulama
                if cell.value in ("VIOLATION!", "NOT COVERED", "NO"):
                    cell.fill = warn_fill
                    cell.font = warn_font
                else:
                    cell.font   = body_font
                    cell.fill   = fill
                cell.border    = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                header_val = ws.cell(row=1, column=cell.column).value or ""
                if "Ratio" in str(header_val) or "Pct" in str(header_val):
                    cell.number_format = "0.00"
                elif isinstance(cell.value, float):
                    cell.number_format = "#,##0.00"

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max(
                (len(str(c.value)) for c in col_cells if c.value is not None),
                default=8
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 48)

        ws.row_dimensions[1].height = 30

    wb.save(output_file)


def export_results(params, model, x, y, q, w):
    tables = build_report_tables(params, model, x, y, q, w)
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        for sheet_name, df in tables.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    format_workbook(OUTPUT_FILE)

    objective   = pl.value(model.objective)
    status      = pl.LpStatus[model.status]
    ext_water   = sum(_val(w[j]) for j in params["J"])

    print("\n" + "=" * 65)
    print("  DÜZELTİLMİŞ MODEL SONUÇLARI")
    print("=" * 65)
    print(f"  Çözüm Durumu   : {status}")
    print(f"  Toplam Maliyet : {objective:,.0f} TL")
    print(f"  Dış Kaynak Su  : {ext_water:.1f} m³")
    print(f"  Çıktı dosyası  : {OUTPUT_FILE}")
    print("=" * 65)


# =============================================================================
# 5) ANA AKIŞ
# =============================================================================

def main():
    if not INPUT_FILE.exists():
        raise FileNotFoundError(
            f"Girdi dosyası bulunamadı: {INPUT_FILE}\n"
            "Python dosyasıyla Excel'i aynı klasöre koyun."
        )
    print("[1/3] Parametreler okunuyor...")
    params = build_parameters()
    print(f"      I={len(params['I'])} talep, J={len(params['J'])} aday lokasyon, K={len(params['K'])} tank tipi")
    print("[2/3] Model çözülüyor (DÜZELTİLMİŞ KISITLARLA)...")
    model, x, y, q, w = solve_model(params)
    print("[3/3] Sonuçlar dışa aktarılıyor...")
    export_results(params, model, x, y, q, w)


if __name__ == "__main__":
    main()